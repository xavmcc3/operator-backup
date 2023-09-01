from multiprocessing import Process, freeze_support
from openpyxl.formula.translate import Translator
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
from openpyxl.styles import NamedStyle
from tkinter import Tk
import pandas as pd
import datetime
import ctypes
import PIL
import re

from clrprint import clrprint
from pathlib import Path
from io import BytesIO
from copy import copy
import subprocess
import openpyxl
import shutil
import sys
import os

# Compile with `python -m PyInstaller --clean operator-tool.spec`
# From `python -m PyInstaller --clean --onefile --icon=operator-tool.ico --name=operator-tool main.py`

hwnd = ctypes.windll.user32.GetForegroundWindow()

def restore_focus():
    ctypes.windll.user32.BringWindowToTop(hwnd)

def file_dialog(filetypes):
    return askopenfilename(filetypes=filetypes)

def folder_dialog():
    return askdirectory()

def empty_dir(path):
    print(f"Emptying ", end="")
    clrprint(path, end="", clr='m')
    print("...", end="\r")

    for filename in os.scandir(path):
        file_path = os.path.join(path, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))

def file_is_valid(file: str):
    path = Path(file)
    if "Template" in path.stem:
        return False
    if "~$" in path.stem:
        return False
    if path.suffix in ['.xlsx']:
        return True
    return False

def copy_row(row, row_index, sheet):
        for cell in row:
            sheet.cell(row=row_index, column=cell.column).value = cell.value

def get_name_from_file(src):
    pattern = r'[0-9]'
    path = Path(src)

    rename = re.sub(pattern, '', path.stem)
    return rename.strip()

def load_path_bytes(path):
    with open(path, 'rb') as f:
        template_bytes = BytesIO(f.read())
    return template_bytes

def wb_from_bytes(path_bytes):
    return openpyxl.load_workbook(path_bytes)

def wb_data_from_bytes(path_bytes):
    return openpyxl.load_workbook(path_bytes, data_only=True)

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)
    copy_sheet_attributes(source_sheet, target_sheet)

def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    target_sheet.column_dimensions = source_sheet.column_dimensions
    target_sheet.row_dimensions = source_sheet.row_dimensions

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

def create_from_template(template_bytes, name, append, dst_folder, use_name=True):
    dst_path = f"{name}{append}.xlsx"
    final_path = os.path.join(os.path.abspath(dst_folder), dst_path)

    print(f"Loading ", end="")
    clrprint(name, end="", clr='y')
    print("...")

    template = wb_from_bytes(template_bytes)

    print(f"Creating ", end="")
    clrprint(dst_path, end="", clr='m')
    print("...")

    ws = template[template.sheetnames[0]]
    for col in ws.iter_cols():
        if str(col[0].value).lower() == "name":
            col[0].value = name if use_name else "Name"
            break
    
    #ws.column_dimensions.group(start='H', end='K', hidden=True)

    template.save(final_path)
    template.close()

    print("Created ", end="")
    clrprint(final_path, clr='g')

def create_all(template_path, src, append, dst):
    print("")
    start_time = datetime.datetime.now()
    processes = []
    template_bytes = load_path_bytes(template_path)
    p = Process(target=create_from_template, args=(template_bytes, "AATemplate", append, dst, False,))
    processes.append(p)
    p.start()
    for file in os.listdir(src):
        if file_is_valid(file):
            name = get_name_from_file(file)
            p = Process(target=create_from_template, args=(template_bytes, name, append, dst,))
            processes.append(p)
            p.start()

    for p in processes:
        p.join()

    clrprint("Finished in ", f"{str(datetime.datetime.now() - start_time)}", clr="w,b")


def archive_before(file_bytes, name, year, dest):
    clrprint("Archiving ", name, " before ", year, "...", sep="", clr="w,y,w,m,w")
    year_date = datetime.datetime(year, 1, 1)

    archive_wb = wb_data_from_bytes(file_bytes)
    archive_ws = archive_wb.get_sheet_by_name(archive_wb.sheetnames[0])
    clrprint("Getting rows in ", name, "...", sep="", clr="w,y,w")

    blank = 0
    row_count = 1
    temp_sheet = archive_wb.create_sheet(f'', index=0)
    for row in archive_ws.iter_rows(min_col=1, max_col=11, min_row=1):
        try:
            date = pd.to_datetime(row[5].value)
        except (IndexError, pd._libs.tslibs.parsing.DateParseError):
            copy_row(row, row_count, temp_sheet)
            row_count += 1
            continue
        if date is None:
            blank += 1
            if blank > 9:
                break
            continue
        if date is None or date > year_date:
            continue
        copy_row(row, row_count, temp_sheet)
        row_count += 1
    
    archive_wb.remove_sheet(archive_ws)
    temp_sheet.title = archive_ws.title

    temp_sheet.column_dimensions = archive_ws.column_dimensions
    temp_sheet.row_dimensions = archive_ws.row_dimensions

    clrprint("Saving ", dest, "...", sep="", clr="w,m,w")
    archive_wb.save(os.path.abspath(dest))
    archive_wb.close()

def remove_before(file_bytes, name, year, dest, temp: openpyxl.Workbook):
    clrprint("Cleaning ", name, " before ", year, "...", sep="", clr="w,y,w,m,w")
    year_date = datetime.datetime(year, 1, 1)

    current_wb = wb_data_from_bytes(file_bytes)
    current_ws = current_wb.get_sheet_by_name(current_wb.sheetnames[0])
    clrprint("Getting rows in ", name, "...", sep="", clr="w,y,w")
    temp_sheet = current_wb.create_sheet(f'', index=0)

    blank = 0
    row_count = 1
    for row in current_ws.iter_rows(min_col=1, max_col=11, min_row=1):
        try:
            date = pd.to_datetime(row[5].value)
        except (IndexError, pd._libs.tslibs.parsing.DateParseError):
            copy_row(row, row_count, temp_sheet)
            row_count += 1
            continue
        if date is None:
            blank += 1
            if blank > 9:
                break
            continue
        if date < year_date:
            continue
        copy_row(row, row_count, temp_sheet)
        row_count += 1

    formulas = {}
    forms_ws = temp.get_sheet_by_name(temp.sheetnames[0])
    for col in forms_ws.iter_cols(min_row=2, max_row=2, min_col=1, max_col=11):
        for cell in col:
            if not str(cell.value).startswith('='):
                continue

            formulas[cell.column] = (cell.value, str(cell.column_letter) + str(cell.row))
    
    min_cap = temp_sheet.max_row
    style = NamedStyle(name='custom_datetime', number_format='MM/DD/YY h:mm AM/PM')
    for row in range(2, temp_sheet.max_row + 1):
        temp_sheet[f"F{row}"].style = style
    for row in range(min_cap + 1, max(5000, (temp_sheet.max_row + 1))):
        for column in formulas.keys():
            cell_idx = f"{openpyxl.utils.cell.get_column_letter(column)}{row}"
            try:
                temp_sheet[cell_idx] = Translator(*formulas[column]).translate_formula(cell_idx)
            except Exception as e:
                print(e)
            if column == 6:
                temp_sheet[cell_idx].style = style

    current_wb.remove_sheet(current_wb.get_sheet_by_name(current_wb.sheetnames[-1]))
    temp_sheet.freeze_panes = temp_sheet['A2']
    current_wb.remove_sheet(current_ws)
    temp_sheet.title = current_ws.title

    stats_ws = temp.get_sheet_by_name(temp.sheetnames[1])
    temp_stats = current_wb.create_sheet(f'', index=1)
    temp_stats.title = stats_ws.title
    copy_sheet(stats_ws, temp_stats)

    temp_sheet.column_dimensions = current_ws.column_dimensions
    temp_sheet.row_dimensions = current_ws.row_dimensions

    temp_stats.column_dimensions = stats_ws.column_dimensions
    temp_stats.row_dimensions = stats_ws.row_dimensions

    clrprint("Saving ", dest, "...", sep="", clr="w,m,w")
    current_wb.save(os.path.abspath(dest))
    current_wb.close()

def extract_years(path, year, dst_folder, temp):
    name = get_name_from_file(path)
    clrprint("Loading ", name, "...", sep="", clr='w,y,w')

    file_bytes = load_path_bytes(path)
    archive_before(file_bytes, name, year, f"{dst_folder}/{name}{year - 1}.xlsx")
    remove_before(file_bytes, name, year, path, temp)

def extract_all(src, year, pth):
    start_time = datetime.datetime.now()
    dst_folder = f'Archive{year - 1}'
    if not os.path.exists(dst_folder):
        os.makedirs(dst_folder)

    clrprint("Loading ", pth, "...", sep="", clr="w,y,w")
    temp_wb = openpyxl.load_workbook(pth)
    
    processes = []
    for file in os.listdir(src):
        if file_is_valid(file):
            p = Process(target=extract_years, args=(os.path.join(src, file), year, dst_folder, temp_wb))
            processes.append(p)
            p.start()

    for p in processes:
        p.join()

    clrprint("Finished in ", f"{str(datetime.datetime.now() - start_time)}", clr="w,b")

if __name__ == "__main__":
    Tk().withdraw()
    freeze_support()
    print("Welcome to ", end="")
    clrprint("Operator Tool", clr='m')

    print("Choose a starting folder")
    src_dir = './data'#folder_dialog()
    print(src_dir)

    if src_dir == "":
        sys.exit()

    # print("Choose a target folder")
    # dst_dir = folder_dialog()
    # print(dst_dir)

    temp_path = "./originals/AATemplate2022.xlsx"

    # if dst_dir == "":
    #     sys.exit()

    # if template == "":
    #     sys.exit()

    try:
        restore_focus()
        year = int('balls')#input("Input a year "))
    except ValueError as e:
        clrprint("Error, defaulting to current year", clr="r")
        year = datetime.date.today().year
    print(year)

    # if True:
    #     empty_dir(dst_dir)

    try:
        result = extract_all(src_dir, year, temp_path)
    except KeyboardInterrupt:
        print("Program terminated, new files might be corrupted")
    #os.startfile(result)
